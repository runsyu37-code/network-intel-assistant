"""
Build SSM Survey Template v2
----------------------------
Rebuilds every sheet of template.xlsx based on:
- SSM_PROJECT.md (hierarchy + Next Phase 3-table split)
- MS_SQL_DRAW.pdf (Susan's hand-drawn SQL design)
- Susan's instructions:
    * Add Site_ID FK to Building/Floor/Room/Rack and all device sheets
    * Add U sub-position (.1 .2 .3) for rack mounting precision
    * Add Created (datetime) + Updated (datetime) to every table
    * Fill Site + Rack sheets (currently empty)
    * Verify CCTV / NVR / Switch sheets against SQL design

Author: Claude (Cowork)  |  May 2026
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
#  STYLE CONSTANTS
# ============================================================
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Color palette by column group
COLORS = {
    "title":     "1F4E78",   # dark blue
    "subtitle":  "B4C7E7",
    "loc":       "DDEBF7",   # location group
    "ident":     "FFF2CC",   # device identity group
    "net":       "E2EFDA",   # network group
    "spec":      "FCE4D6",   # spec / storage
    "ts":        "EDEDED",   # timestamps (Created/Updated)
    "key":       "FFE699",   # PK / FK columns
    "example":   "F2F2F2",
}

def hdr_fill(color):  return PatternFill("solid", fgColor=color)


def apply_title(ws, text, ncols, row=1):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = hdr_fill(COLORS["title"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def apply_subtitle(ws, text, ncols, row):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    c.font = Font(italic=True, size=10, color="1F4E78")
    c.fill = hdr_fill(COLORS["subtitle"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def apply_header_row(ws, headers, row, color_map):
    """
    headers     = list of header texts
    color_map   = list (same length) — key into COLORS (group) per column
    """
    for col_idx, (text, grp) in enumerate(zip(headers, color_map), start=1):
        c = ws.cell(row=row, column=col_idx, value=text)
        c.font = Font(bold=True, size=10, color="000000")
        c.fill = hdr_fill(COLORS[grp])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 42
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def apply_example_row(ws, example_values, row):
    for col_idx, v in enumerate(example_values, start=1):
        c = ws.cell(row=row, column=col_idx, value=v)
        c.font = Font(italic=True, color="7F7F7F", size=9)
        c.fill = hdr_fill(COLORS["example"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def add_blank_rows(ws, start_row, n=20, ncols=None):
    """Add empty rows with borders for data entry."""
    for r in range(start_row, start_row + n):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c, value=None)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def add_dv(ws, options, col_letter, row_start, row_end, prompt=None):
    """Attach a list data-validation dropdown to a column range."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
    )
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)


# ============================================================
#  BUILD WORKBOOK
# ============================================================
wb = Workbook()

# remove default sheet
wb.remove(wb.active)


# ============================================================
#  SHEET 0 — README / INSTRUCTIONS
# ============================================================
ws = wb.create_sheet("📋 README")
set_col_widths(ws, [4, 28, 80])
apply_title(ws, "SSM Survey Template v2 — README", 3, row=1)
apply_subtitle(ws, "How to fill this template for import into MS SQL Server (SSM_DB)", 3, row=2)

readme_rows = [
    ("", "Purpose",
     "Capture every Site → Building → Floor → Room → Rack → Device for SSM v1.0 monitoring."),
    ("", "Order of filling",
     "1) Site → 2) Building → 3) Floor → 4) Room → 5) Rack → 6) CCTV / NVR / SW.  "
     "Always fill parents BEFORE children so FK resolution works."),
    ("", "Site_ID column",
     "Every sheet from Building onward carries Site_ID. Company has multiple sites and "
     "buildings can share names across sites — Site_ID disambiguates them."),
    ("", "U Position",
     "Two columns: u_position (1..total_units) + u_subposition (1..3). "
     "Each U is divided into 3 micro-slots; leave u_subposition blank if device sits flush on the U boundary."),
    ("", "Timestamps",
     "Leave Created / Updated blank — the Python import script fills them with the import time. "
     "If you set Updated manually, the importer respects your value."),
    ("", "Image fields",
     "Put the relative path to the image (e.g. images/floor_b1.png). "
     "Importer will load the file, base64-encode it, and store into image_data + image_type."),
    ("", "Status values",
     "online / offline / warning / unknown (lowercase). Dropdown is enforced where applicable."),
    ("", "Device sync (CCTV/NVR/SW)",
     "Device tables are linked through ip_address, serial_no, and mac_address. "
     "Importer will WARN on duplicate serial_no or invalid IP format."),
    ("", "Roles",
     "Users table seeded with: admin (full edit), user (Rack/Room edit), viewer (read-only, no password). "
     "Password column stores bcrypt hash — leave plain text and importer will hash it."),
]
for i, (a, b, c) in enumerate(readme_rows, start=4):
    ws.cell(row=i, column=2, value=b).font = Font(bold=True)
    ws.cell(row=i, column=2).alignment = Alignment(vertical="top")
    ws.cell(row=i, column=3, value=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 44


# ============================================================
#  SHEET 1 — SITE
# ============================================================
ws = wb.create_sheet("1_Site")
ncols = 8
apply_title(ws, "Layer 1 — Sites (สาขา / สำนักงาน)", ncols, 1)
apply_subtitle(ws,
    "1 แถวต่อ 1 สาขา — Site_ID จะถูกใช้อ้างอิงในทุก sheet ถัดไป", ncols, 2)

headers = ["#", "Site_ID (PK)", "Site Name", "Site Code", "Location / Address",
           "Description", "Created", "Updated"]
groups  = ["key", "key", "ident", "ident", "loc", "spec", "ts", "ts"]
apply_header_row(ws, headers, 3, groups)

example = ["e.g.", "HQ", "สำนักงานใหญ่", "HQ", "กรุงเทพฯ",
           "Main office — 5 buildings", "(auto)", "(auto)"]
apply_example_row(ws, example, 4)
set_col_widths(ws, [5, 14, 24, 12, 32, 32, 18, 18])
add_blank_rows(ws, 5, 25, ncols)


# ============================================================
#  SHEET 2 — BUILDING
# ============================================================
ws = wb.create_sheet("2_Building")
ncols = 11
apply_title(ws, "Layer 2 — Buildings (อาคาร)", ncols, 1)
apply_subtitle(ws,
    "1 แถวต่อ 1 อาคาร — ต้องระบุ Site_ID ให้ตรงกับ sheet 1_Site", ncols, 2)

headers = ["#", "Site_ID (FK)", "Building_ID (PK)", "Building Name",
           "Building Code", "Floor Count", "Description",
           "Image Path", "Note", "Created", "Updated"]
groups  = ["key", "key", "key", "ident", "ident", "spec", "spec",
           "spec", "spec", "ts", "ts"]
apply_header_row(ws, headers, 3, groups)
example = ["e.g.", "HQ", "BLD_A", "อาคาร A", "BLD-A", 8,
           "Main tower", "images/bld_a.png", "ใกล้ลานจอด", "(auto)", "(auto)"]
apply_example_row(ws, example, 4)
set_col_widths(ws, [5, 12, 14, 22, 12, 11, 26, 22, 18, 18, 18])
add_blank_rows(ws, 5, 25, ncols)


# ============================================================
#  SHEET 3 — FLOOR
# ============================================================
ws = wb.create_sheet("3_Floor")
ncols = 12
apply_title(ws, "Layer 3 — Floors (ชั้น)", ncols, 1)
apply_subtitle(ws,
    "1 แถวต่อ 1 ชั้น — Site_ID + Building_ID ใช้สำหรับ sync ขึ้นไปยัง parent", ncols, 2)

headers = ["#", "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (PK)",
           "Floor Number", "Floor Name", "Main Function",
           "Has CCTV?", "Image Path", "Note", "Created", "Updated"]
groups  = ["key", "key", "key", "key", "ident", "ident", "spec",
           "spec", "spec", "spec", "ts", "ts"]
apply_header_row(ws, headers, 3, groups)
example = ["e.g.", "HQ", "BLD_A", "BLD_A_F-1", -1, "B1 (Basement)",
           "Server Room", "Yes", "images/bld_a_b1.png", "—",
           "(auto)", "(auto)"]
apply_example_row(ws, example, 4)
set_col_widths(ws, [5, 12, 14, 16, 12, 18, 18, 12, 22, 18, 18, 18])
add_blank_rows(ws, 5, 30, ncols)

# Has CCTV dropdown
add_dv(ws, ["Yes", "No"], "H", 5, 50, prompt="Yes / No")


# ============================================================
#  SHEET 4 — ROOM
# ============================================================
ws = wb.create_sheet("4_Room")
ncols = 14
apply_title(ws, "Layer 4 — Rooms (ห้อง)", ncols, 1)
apply_subtitle(ws,
    "1 แถวต่อ 1 ห้อง — Site_ID เผื่อ Building name ซ้ำข้ามสาขา", ncols, 2)

headers = ["#", "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (FK)",
           "Room_ID (PK)", "Room Name", "Room Type",
           "Has NVR", "Has SW", "Image Path", "Note",
           "Map_x", "Map_y", "Created"]
# Drop Updated from the visible columns since space tight — handled by importer
# Actually let's add it back; remove Map fields → those go in CCTV table
headers = ["#", "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (FK)",
           "Room_ID (PK)", "Room Name", "Room Type",
           "Has NVR", "Has SW", "Image Path", "Note",
           "Created", "Updated"]
groups  = ["key", "key", "key", "key", "key", "ident", "ident",
           "spec", "spec", "spec", "spec", "ts", "ts"]
ncols = len(headers)
apply_header_row(ws, headers, 3, groups)
example = ["e.g.", "HQ", "BLD_A", "BLD_A_F-1", "BLD_A_B1_SRV01",
           "Server Room 01", "Server Room",
           "Yes", "Yes", "images/srv01.png",
           "เครื่องปรับอากาศ 2 ตัว", "(auto)", "(auto)"]
apply_example_row(ws, example, 4)
set_col_widths(ws, [5, 12, 14, 14, 18, 18, 14, 10, 10, 22, 22, 18, 18])
add_blank_rows(ws, 5, 30, ncols)

add_dv(ws, ["server", "network", "office", "power"], "G", 5, 50,
       prompt="ประเภทห้อง")
add_dv(ws, ["Yes", "No"], "H", 5, 50)
add_dv(ws, ["Yes", "No"], "I", 5, 50)


# ============================================================
#  SHEET 5 — RACK
# ============================================================
ws = wb.create_sheet("5_Rack")
ncols = 14
apply_title(ws, "Layer 5 — Racks (ตู้ Rack)", ncols, 1)
apply_subtitle(ws,
    "1 แถวต่อ 1 Rack — total_units = ความสูงของ Rack ในหน่วย U (เช่น 9, 12, 24, 42)",
    ncols, 2)

headers = ["#", "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (FK)",
           "Room_ID (FK)", "Rack_ID (PK)", "Rack Name",
           "Total U", "Units per U (slots)", "Position X (m)", "Position Y (m)",
           "Note", "Created", "Updated"]
groups  = ["key", "key", "key", "key", "key", "key", "ident",
           "spec", "spec", "loc", "loc", "spec", "ts", "ts"]
apply_header_row(ws, headers, 3, groups)
example = ["e.g.", "HQ", "BLD_A", "BLD_A_F-1", "BLD_A_B1_SRV01",
           "RACK_01", "Rack 01 (Left wall)", 9, 3, 1.5, 2.0,
           "Standard 9U cabinet", "(auto)", "(auto)"]
apply_example_row(ws, example, 4)
set_col_widths(ws, [5, 12, 14, 14, 18, 14, 22, 9, 12, 12, 12, 24, 18, 18])
add_blank_rows(ws, 5, 30, ncols)


# ============================================================
#  HELPER: build a device-style sheet (CCTV / NVR / Switch)
# ============================================================
def build_device_sheet(ws, title, sub, headers, groups, example, widths,
                       status_col_letter=None, dropdowns=None):
    ncols = len(headers)
    apply_title(ws, title, ncols, 1)
    apply_subtitle(ws, sub, ncols, 2)
    apply_header_row(ws, headers, 3, groups)
    apply_example_row(ws, example, 4)
    set_col_widths(ws, widths)
    add_blank_rows(ws, 5, 40, ncols)
    if status_col_letter:
        add_dv(ws,
               ["online", "offline", "warning", "unknown"],
               status_col_letter, 5, 60,
               prompt="ค่าที่ระบบใช้ติดตาม")
    if dropdowns:
        for col, opts, prompt in dropdowns:
            add_dv(ws, opts, col, 5, 60, prompt=prompt)


# ============================================================
#  SHEET 6 — CCTV (cameras)
# ============================================================
ws = wb.create_sheet("6_CCTV")
headers = [
    "#",
    "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (FK)", "Room_ID (FK)", "Rack_ID (FK opt.)",
    # U position
    "U Position", "U Sub-pos (1-3)",
    # Identity (sync keys)
    "Device Name", "Brand", "Model", "Serial No (S/N)", "MAC Address",
    "Camera Type", "Resolution",
    # Network
    "IP Address", "VLAN",
    # PoE Switch link
    "PoE Switch Name", "Switch Port",
    # NVR link
    "NVR Device Name", "NVR Channel",
    # Map position on room SVG
    "Map_x", "Map_y",
    # Status
    "Status", "Fail Count", "Note",
    # Timestamps
    "Created", "Updated",
]
groups = [
    "key",
    "key","key","key","key","key",
    "loc","loc",
    "ident","ident","ident","ident","ident","ident","ident",
    "net","net",
    "net","net",
    "spec","spec",
    "loc","loc",
    "spec","spec","spec",
    "ts","ts",
]
example = [
    "e.g.",
    "HQ","BLD_A","BLD_A_F-1","BLD_A_B1_SRV01","RACK_01",
    2, 1,
    "A-B1-SRV-CAM01","Hikvision","DS-2CD2143G2-I","SN123456","AA:BB:CC:11:22:33",
    "Dome","4MP",
    "192.168.1.101","VLAN10",
    "SW-B1-01", 1,
    "NVR-B1-01", 1,
    120, 80,
    "online", 0, "ติดเพดานทางเข้า",
    "(auto)","(auto)",
]
widths = [5, 11, 13, 13, 16, 13, 9, 11,
          22, 12, 22, 18, 18, 12, 12,
          16, 9, 16, 11, 16, 11, 9, 9,
          11, 10, 28, 18, 18]
build_device_sheet(
    ws,
    "Layer 6 — CCTV Cameras",
    "★ S/N + IP + MAC = sync keys  |  U Position + Sub-pos ระบุตำแหน่งในตู้ Rack",
    headers, groups, example, widths,
    status_col_letter="X",
)


# ============================================================
#  SHEET 7 — NVR
# ============================================================
ws = wb.create_sheet("7_NVR")
headers = [
    "#",
    "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (FK)", "Room_ID (FK)", "Rack_ID (FK)",
    "U Position", "U Sub-pos (1-3)",
    "Device Name", "Brand", "Model", "Serial No (S/N)", "MAC Address",
    "OS / Firmware", "VLAN",
    "IP (Internet Port)", "IP (CCTV Port)",
    "Total Channels", "Active Channels",
    "HDD Total (TB)", "HDD Used (%)",
    "Recording Resolution", "Retention (days)", "Record Status",
    "Status", "Fail Count", "Note",
    "Created", "Updated",
]
groups = [
    "key",
    "key","key","key","key","key",
    "loc","loc",
    "ident","ident","ident","ident","ident",
    "ident","net",
    "net","net",
    "spec","spec",
    "spec","spec",
    "spec","spec","spec",
    "spec","spec","spec",
    "ts","ts",
]
example = [
    "e.g.",
    "HQ","BLD_A","BLD_A_F-1","BLD_A_B1_SRV01","RACK_01",
    1, 1,
    "NVR-B1-01","Hikvision","DS-9632NI-I8","SN999111","AA:BB:CC:99:99:99",
    "v4.71.005","VLAN10",
    "10.0.0.50","192.168.1.200",
    32, 28,
    16, 65,
    "1080p", 30, "normal",
    "online", 0, "—",
    "(auto)","(auto)",
]
widths = [5, 11, 13, 13, 16, 13, 9, 11,
          18, 12, 22, 18, 18, 14, 9,
          15, 15, 11, 11,
          10, 10, 16, 11, 12,
          11, 10, 22, 18, 18]
build_device_sheet(
    ws,
    "Layer 6 — NVR (Network Video Recorder)",
    "★ HDD Used > 80% → ระบบจะแจ้งเตือน  |  Record Status: normal / warning / fail",
    headers, groups, example, widths,
    status_col_letter="Y",
    dropdowns=[
        ("X", ["normal", "warning", "fail", "unknown"], "สถานะการบันทึก"),
    ],
)


# ============================================================
#  SHEET 8 — Switch (PoE / Net)
# ============================================================
ws = wb.create_sheet("8_Switch")
headers = [
    "#",
    "Site_ID (FK)", "Building_ID (FK)", "Floor_ID (FK)", "Room_ID (FK)", "Rack_ID (FK)",
    "U Position", "U Sub-pos (1-3)",
    "Device Name", "Switch Type",
    "Brand", "Model", "Serial No (S/N)", "MAC Address",
    "OS / Firmware", "VLAN",
    "IP Address",
    "Total Ports", "PoE Ports",
    "PoE Budget (W)", "PoE Used (W)",
    "Uplink Port", "Status", "Fail Count", "Note",
    "Created", "Updated",
]
groups = [
    "key",
    "key","key","key","key","key",
    "loc","loc",
    "ident","ident",
    "ident","ident","ident","ident",
    "ident","net",
    "net",
    "spec","spec",
    "spec","spec",
    "net","spec","spec","spec",
    "ts","ts",
]
example = [
    "e.g.",
    "HQ","BLD_A","BLD_A_F-1","BLD_A_B1_SRV01","RACK_01",
    5, 1,
    "SW-B1-01","PoE",
    "Zyxel","GS2220-50HP","SN777","AA:BB:CC:77:77:77",
    "v4.80","VLAN10",
    "192.168.1.1",
    50, 48,
    375, 180,
    "Core-SW", "online", 0, "—",
    "(auto)","(auto)",
]
widths = [5, 11, 13, 13, 16, 13, 9, 11,
          18, 12,
          12, 22, 18, 18,
          14, 9,
          15,
          11, 10, 12, 12,
          14, 11, 10, 22, 18, 18]
build_device_sheet(
    ws,
    "Layer 6 — PoE / Network Switch",
    "★ Switch Type: PoE / Non-PoE / Core  |  PoE Budget เกิน 80% = warning",
    headers, groups, example, widths,
    status_col_letter="W",
    dropdowns=[
        ("J", ["PoE", "Non-PoE", "Core", "Aggregation"], "ประเภท switch"),
    ],
)


# ============================================================
#  SHEET 9 — Users (login + roles)
# ============================================================
ws = wb.create_sheet("9_Users")
ncols = 10
apply_title(ws, "Auth — Users / Login (3 roles)", ncols, 1)
apply_subtitle(ws,
    "admin = แก้ได้ทั้งหมด  |  user = แก้ Rack/Room เท่านั้น  |  viewer = อ่านอย่างเดียว ไม่ต้องใส่รหัส",
    ncols, 2)

headers = ["#", "User_ID (PK)", "Username", "Password (plain → จะถูก hash)",
           "Display Name", "Role", "Is Active",
           "Last Login", "Created", "Updated"]
groups  = ["key", "key", "ident", "ident", "ident", "spec", "spec",
           "ts", "ts", "ts"]
apply_header_row(ws, headers, 3, groups)
example = ["e.g.", 1, "admin", "12345", "System Admin", "admin", "Yes",
           "(auto)", "(auto)", "(auto)"]
apply_example_row(ws, example, 4)
set_col_widths(ws, [5, 13, 16, 26, 22, 12, 11, 18, 18, 18])
add_blank_rows(ws, 5, 10, ncols)

add_dv(ws, ["admin", "user", "viewer"], "F", 5, 14,
       prompt="3 roles: admin / user / viewer")
add_dv(ws, ["Yes", "No"], "G", 5, 14)


# ============================================================
#  SAVE
# ============================================================
out_path = "/sessions/festive-modest-hawking/mnt/outputs/template_v2.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
