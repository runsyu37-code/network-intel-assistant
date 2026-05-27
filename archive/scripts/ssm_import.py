#!/usr/bin/env python3
"""
ssm_import.py — Import template_v3.xlsx into MS SQL Server (SSM_DB).

Install deps:
    pip install openpyxl pyodbc bcrypt
    (also needs: Microsoft ODBC Driver 17+ for SQL Server)

Usage:
    # Windows Auth — home laptop
    python ssm_import.py template_v3.xlsx --server localhost\\SQLEXPRESS --db SSM_DB --auth windows

    # SQL Auth — work notebook
    python ssm_import.py template_v3.xlsx --server SRV\\SQLEXPRESS --db SSM_DB --auth sql --user sa --password MyPass1

    # Dry run: reads Excel + validates, prints what would be inserted, then ROLLS BACK
    python ssm_import.py template_v3.xlsx --server localhost\\SQLEXPRESS --db SSM_DB --auth windows --dry-run

    # Parse only: no DB needed — just validates Excel and shows rows
    python ssm_import.py template_v3.xlsx --parse-only
"""

import argparse
import base64
import csv
import mimetypes
import re
import sys
from pathlib import Path

# ── Dependency checks ─────────────────────────────────────────────────────────

try:
    import openpyxl
except ImportError:
    sys.exit("Missing: pip install openpyxl")

try:
    import pyodbc
except ImportError:
    sys.exit("Missing: pip install pyodbc  (also needs ODBC Driver 17+ for SQL Server)")

try:
    import bcrypt as _bcrypt
except ImportError:
    sys.exit("Missing: pip install bcrypt")

# ── Regex validators ──────────────────────────────────────────────────────────

IP_RE  = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
MAC_RE = re.compile(r"^([0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}$")
MAX_IMAGE_BYTES = 5 * 1024 * 1024  # 5 MB hard limit (D10)

# ── Sheet -> Table config (MUST be in FK dependency order) ────────────────────

SHEETS = [
    dict(sheet="1_Site",     table="sites",        pk="Site_ID"),
    dict(sheet="2_Building", table="buildings",    pk="Building_ID"),
    dict(sheet="3_Floor",    table="floors",       pk="Floor_ID"),
    dict(sheet="4_Room",     table="rooms",        pk="Room_ID"),
    dict(sheet="5_Rack",     table="racks",        pk="Rack_ID"),
    dict(sheet="8_Switch",   table="poe_switches", pk="SW_ID"),
    dict(sheet="7_NVR",      table="nvrs",         pk="NVR_ID"),
    dict(sheet="6_CCTV",     table="cameras",      pk=None),   # IDENTITY PK
    # dict(sheet="9_Users",    table="users",        pk=None),   # skip: seeded by schema already
]

# ── Header alias table (lowercase Excel header -> SQL column name) ─────────────
# None = skip this column entirely

ALIAS: dict[str, str | None] = {
    # PK / FK annotations — strip the suffix, keep the ID
    "site_id (pk)":          "Site_ID",
    "building_id (pk)":      "Building_ID",
    "floor_id (pk)":         "Floor_ID",
    "room_id (pk)":          "Room_ID",
    "rack_id (pk)":          "Rack_ID",
    "nvr_id (pk)":           "NVR_ID",
    "sw_id (pk)":            "SW_ID",           # v3
    "sw_id(pk)":             "SW_ID",           # v4 (no space)
    "nvr_ch (pk)":           "NVR_CH",          # v3
    "id(pk)":                "NVR_CH",          # v4 CCTV
    "user_id (pk)":          None,           # IDENTITY — SQL Server assigns
    "site_id (fk)":          "Site_ID",
    "building_id (fk)":      "Building_ID",
    "floor_id (fk)":         "Floor_ID",
    "room_id (fk)":          "Room_ID",
    "rack_id (fk)":          "Rack_ID",
    "nvr_id (fk)":           "NVR_ID",
    "sw_id (fk)":            "SW_ID",
    # name columns
    "site name":             "name",
    "building name":         "name",
    "floor name":            "name",
    "rack name":             "name",
    "room name":             "name",
    "device name":           "device_name",
    "display name":          "display_name",
    # hierarchy details
    "site code":             "code",
    "building code":         "code",
    "location / address":    "location",
    "floor number":          "floor_number",
    "floor count":           "floor_count",
    "main function":         "function",
    "has cctv?":             "has_cctv",
    "has nvr":               "has_nvr",
    "has sw":                "has_sw",
    "room type":             "type",
    "total u":               "total_units",
    "units per u":           "units_per_u",
    # device identity
    "brand":                 "brand",
    "model":                 "model",
    "s/n":                   "serial_no",       # v3
    "serial no (s/n)":       "serial_no",       # v4
    "mac":                   "mac_address",     # v3
    "mac address":           "mac_address",     # v4
    "camera type":           "camera_type",
    "resolution":            "resolution",
    "switch type":           "switch_type",
    "os/firmware":           "os_version",      # v3
    "os / firmware":         "os_version",      # v4
    # network
    "ip address":            "ip_address",
    "ip (internet port)":    "ip_internet",
    "ip (cctv port)":        "ip_cctv",
    "vlan":                  "vlan_id",
    "subnet mask":           "subnet_mask",
    "gateway":               "gateway",
    # rack position (D2)
    "u position":            "u_position",
    "u sub-pos":             "u_subposition",
    "u-size":                "u_size",
    # switch-specific
    "total ports":           "total_ports",
    "poe ports":             "poe_ports",
    "poe budget (w)":        "poe_budget_w",
    "poe used (w)":          "poe_used_w",
    "uplink port":           "uplink_port",
    # camera-specific (D12: SW_ID + NVR_ID typed directly)
    "poe switch name":       "SW_ID",
    "switch port":           "poe_port_number",
    "nvr device name":       "NVR_ID",
    "nvr":                   "NVR_ID",          # v4 short form
    "nvr channel":           "nvr_channel",
    "ch(port)":              "nvr_channel",     # v4
    "install location":      "install_location",
    # nvr-specific
    "total channels":        "total_channels",
    "active channels":       "active_channels",
    "hdd total (tb)":        "hdd_total_tb",
    "recording resolution":  "recording_res",
    "retention (days)":      "retention_days",
    "record status":         "record_status",
    # auth
    "username":              "username",
    "password (plain)":      "_password",    # v3
    "password (plain → จะถูก hash)": "_password",  # v4
    "role":                  "role",
    "is active":             "is_active",
    # shared
    "description":           "description",
    "status":                "status",
    "fail count":            "fail_count",
    "note":                  "_note",        # resolved per-table (note vs notes)
    "created":               "created_at",
    "updated":               "updated_at",
    "image path":            "_image_path",  # special: base64 encode file
    "#":                     "__row_num",    # row-number column — used for e.g. detection, not inserted
}

# "note" column name differs by table
_NOTE_COL: dict[str, str | None] = {
    "sites":        None,
    "buildings":    "note",
    "floors":       "note",
    "rooms":        "note",
    "racks":        "note",
    "poe_switches": "notes",
    "nvrs":         "notes",
    "cameras":      "notes",
    "users":        None,
}

# BIT columns — Excel has Yes/No/True/False -> 1/0
_BIT_COLS = {"has_cctv", "has_nvr", "has_sw", "is_active"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean(val) -> str | None:
    """Strip whitespace; treat empty / '(auto)' / 'e.g.' as None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "(auto)", "e.g.", "-", "N/A"):
        return None
    return s


def _bit(val) -> int | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("1", "yes", "true", "y"):
        return 1
    if s in ("0", "no", "false", "n"):
        return 0
    return None


def _hash_password(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode(), _bcrypt.gensalt(rounds=12)).decode()


def _encode_image(path_str: str, xlsx_dir: Path) -> tuple[str, str] | tuple[None, None]:
    """Read image file -> (base64 string, mime type) or (None, None) on error."""
    p = Path(path_str)
    if not p.is_absolute():
        p = xlsx_dir / p
    if not p.exists():
        return None, None
    if p.stat().st_size > MAX_IMAGE_BYTES:
        return None, None
    mime, _ = mimetypes.guess_type(str(p))
    mime = mime or "application/octet-stream"
    data = base64.b64encode(p.read_bytes()).decode()
    return data, mime


def _validate_ip(val: str | None) -> str | None:
    if not val:
        return None
    if not IP_RE.match(val):
        raise ValueError(f"Invalid IP: {val!r}")
    return val


def _validate_mac(val: str | None) -> str | None:
    if not val:
        return None
    if not MAC_RE.match(val):
        raise ValueError(f"Invalid MAC (expected XX:XX:XX:XX:XX:XX): {val!r}")
    return val


# ── Sheet reader ──────────────────────────────────────────────────────────────

def _find_header_row(ws) -> int | None:
    """Find the row whose first non-empty cell is '#'."""
    for row in ws.iter_rows():
        for cell in row:
            if _clean(cell.value) == "#":
                return cell.row
    return None


def _build_header_map(ws, header_row: int, table: str) -> dict[int, str | None]:
    """Return {col_index: sql_column_name} for the header row."""
    mapping: dict[int, str | None] = {}
    for cell in ws[header_row]:
        raw = _clean(cell.value)
        if not raw:
            continue
        key = raw.lower()
        if key in ALIAS:
            sql_col = ALIAS[key]
            if sql_col == "_note":
                sql_col = _NOTE_COL.get(table)
            mapping[cell.column] = sql_col
        else:
            mapping[cell.column] = None  # unknown header -> skip
    return mapping


def _parse_row(ws_row, header_map: dict[int, str | None],
               table: str, xlsx_dir: Path) -> dict | None:
    """
    Convert one worksheet row to a {sql_col: value} dict.
    Returns None if the row should be skipped entirely.
    """
    raw: dict[str, object] = {}
    for cell in ws_row:
        sql_col = header_map.get(cell.column)
        if sql_col is None:
            continue
        raw[sql_col] = cell.value

    # Skip example row (check raw value before _clean strips "e.g." to None)
    if str(raw.get("__row_num", "")).strip() == "e.g.":
        return None

    # Special columns
    row: dict[str, object] = {}

    for col, val in raw.items():
        v = _clean(val)

        if col == "_image_path":
            if v:
                img_data, img_type = _encode_image(v, xlsx_dir)
                if img_data:
                    row["image_data"] = img_data
                    row["image_type"] = img_type
            continue

        if col == "_password":
            if v:
                row["pw_hash"] = _hash_password(v)
            continue

        if col.startswith("__"):
            continue

        if col in ("created_at", "updated_at"):
            row[col] = None  # let DB fill with SYSUTCDATETIME()
            continue

        if col in _BIT_COLS:
            row[col] = _bit(v)
            continue

        row[col] = v if v is not None else None

    return row if row else None


# ── Validation ────────────────────────────────────────────────────────────────

def _validate_row(row: dict, table: str) -> tuple[list[str], list[str]]:
    """Return (hard_errors, warnings).
    Warnings nullify the offending field and allow the row to insert.
    Hard errors block the row entirely.
    """
    errors: list[str] = []
    warnings: list[str] = []

    for field in ("ip_address", "ip_internet", "ip_cctv"):
        val = row.get(field)
        if val and not IP_RE.match(str(val)):
            warnings.append(f"Invalid IP in {field}: {val!r} -> NULL")
            row[field] = None

    val = row.get("mac_address")
    if val and not MAC_RE.match(str(val)):
        warnings.append(f"Invalid MAC: {val!r} -> NULL")
        row["mac_address"] = None

    hdd = row.get("hdd_used_pct")
    if hdd is not None:
        try:
            if not (0 <= float(hdd) <= 100):
                errors.append(f"hdd_used_pct out of range: {hdd}")
        except (TypeError, ValueError):
            errors.append(f"hdd_used_pct not a number: {hdd}")

    usub = row.get("u_subposition")
    if usub is not None:
        try:
            if int(usub) not in (1, 2, 3):
                errors.append(f"u_subposition must be 1/2/3, got {usub}")
        except (TypeError, ValueError):
            errors.append(f"u_subposition not an integer: {usub}")

    return errors, warnings


# ── DB helpers ────────────────────────────────────────────────────────────────

def _connect(args) -> "pyodbc.Connection":
    if args.auth == "windows":
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={args.server};DATABASE={args.db};Trusted_Connection=yes;"
        )
    else:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={args.server};DATABASE={args.db};"
            f"UID={args.user};PWD={args.password};"
        )
    return pyodbc.connect(conn_str, autocommit=False)


def _insert(cursor, table: str, row: dict) -> None:
    cols = [c for c, v in row.items() if v is not None]
    if not cols:
        return
    placeholders = ", ".join("?" for _ in cols)
    col_str = ", ".join(f"[{c}]" for c in cols)
    sql = f"INSERT INTO {table} ({col_str}) VALUES ({placeholders})"
    cursor.execute(sql, [row[c] for c in cols])


def _upsert(cursor, table: str, row: dict, pk: str | None) -> str:
    """MERGE: INSERT ถ้าไม่มี, UPDATE ถ้ามีอยู่แล้ว. คืน 'inserted'|'updated'|'skipped'."""
    cols = [c for c, v in row.items() if v is not None]
    if not cols:
        return "skipped"

    # ถ้าไม่มี PK ธรรมชาติ (IDENTITY) → INSERT ธรรมดา
    if not pk or pk not in row or not row[pk]:
        _insert(cursor, table, row)
        return "inserted"

    update_cols = [c for c in cols if c != pk]
    src_select  = ", ".join(f"? AS [{c}]" for c in cols)
    ins_cols    = ", ".join(f"[{c}]" for c in cols)
    ins_vals    = ", ".join(f"source.[{c}]" for c in cols)
    upd_set     = ", ".join(f"target.[{c}] = source.[{c}]" for c in update_cols)

    if update_cols:
        sql = f"""
        MERGE INTO [{table}] WITH (HOLDLOCK) AS target
        USING (SELECT {src_select}) AS source
        ON target.[{pk}] = source.[{pk}]
        WHEN MATCHED THEN UPDATE SET {upd_set}
        WHEN NOT MATCHED THEN INSERT ({ins_cols}) VALUES ({ins_vals})
        OUTPUT $action;
        """
    else:
        sql = f"""
        MERGE INTO [{table}] WITH (HOLDLOCK) AS target
        USING (SELECT {src_select}) AS source
        ON target.[{pk}] = source.[{pk}]
        WHEN NOT MATCHED THEN INSERT ({ins_cols}) VALUES ({ins_vals})
        OUTPUT $action;
        """

    cursor.execute(sql, [row[c] for c in cols])
    result = cursor.fetchone()
    return result[0].lower() if result else "inserted"


# ── Main import logic ─────────────────────────────────────────────────────────

def _import_sheet(cursor, conf: dict, ws, xlsx_dir: Path,
                  report_rows: list, parse_only: bool, upsert: bool = False) -> int:
    table = conf["table"]
    pk    = conf["pk"]

    header_row_num = _find_header_row(ws)
    if header_row_num is None:
        print(f"  [SKIP] {ws.title}: no header row found")
        return 0

    header_map = _build_header_map(ws, header_row_num, table)
    inserted = 0

    for ws_row in ws.iter_rows(min_row=header_row_num + 1):
        # Skip blank rows
        values = [c.value for c in ws_row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        row = _parse_row(ws_row, header_map, table, xlsx_dir)
        if row is None:
            continue

        # Skip if PK is empty (mandatory)
        if pk and not row.get(pk):
            continue

        # cameras: auto-fill device_name from NVR_CH if blank
        if table == "cameras" and not row.get("device_name"):
            row["device_name"] = row.get("NVR_CH") or f"CAM_{inserted + 1}"
        # switches: auto-fill device_name from SW_ID if blank
        if table == "poe_switches" and not row.get("device_name"):
            row["device_name"] = row.get("SW_ID") or f"SW_{inserted + 1}"

        errors, warnings = _validate_row(row, table)
        pk_val = row.get(pk, "(auto)")

        if errors:
            msg = "; ".join(errors)
            print(f"  [FAIL] {table} PK={pk_val}: {msg}")
            report_rows.append({"sheet": ws.title, "pk": pk_val,
                                 "result": "FAIL", "reason": msg})
            continue

        for w in warnings:
            print(f"  [WARN] {table} PK={pk_val}: {w}")

        if parse_only:
            print(f"  [DRY]  {table} PK={pk_val}  cols={list(row.keys())}")
            report_rows.append({"sheet": ws.title, "pk": pk_val,
                                 "result": "DRY", "reason": "; ".join(warnings)})
        elif upsert:
            action = _upsert(cursor, table, row, pk)
            label = "[NEW]" if action == "inserted" else "[UPD]"
            print(f"  {label}  {table} PK={pk_val}")
            report_rows.append({"sheet": ws.title, "pk": pk_val,
                                 "result": action.upper(), "reason": "; ".join(warnings)})
        else:
            _insert(cursor, table, row)
            print(f"  [OK]   {table} PK={pk_val}")
            report_rows.append({"sheet": ws.title, "pk": pk_val,
                                 "result": "OK", "reason": "; ".join(warnings)})
        inserted += 1

    return inserted


def run(args) -> None:
    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    print(f"Opening workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    xlsx_dir = xlsx_path.parent

    report_rows: list[dict] = []
    total = 0

    # ── parse-only mode: no DB ────────────────────────────────────────────────
    if args.parse_only:
        print("Mode: PARSE ONLY (no database connection)\n")
        for conf in SHEETS:
            sheet_name = conf["sheet"]
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if ws is None:
                print(f"[MISSING SHEET] {sheet_name}")
                continue
            print(f"Sheet: {sheet_name} -> {conf['table']}")
            n = _import_sheet(None, conf, ws, xlsx_dir, report_rows,
                              parse_only=True)
            total += n
            print(f"  -> {n} rows parsed\n")
        _write_report(report_rows, args.report)
        print(f"Done. {total} rows parsed (no DB touched).")
        return

    # ── live or dry-run mode: needs DB ────────────────────────────────────────
    mode_label = "UPSERT" if args.upsert else "INSERT"
    print(f"Connecting to {args.server}/{args.db} ({args.auth} auth)... [{mode_label} mode]")
    conn = _connect(args)
    cursor = conn.cursor()

    try:
        for conf in SHEETS:
            sheet_name = conf["sheet"]
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if ws is None:
                print(f"[MISSING SHEET] {sheet_name}")
                continue
            print(f"Sheet: {sheet_name} -> {conf['table']}")
            n = _import_sheet(cursor, conf, ws, xlsx_dir, report_rows,
                              parse_only=False, upsert=args.upsert)
            total += n
            verb = "upserted" if args.upsert else "inserted"
            print(f"  -> {n} rows {verb}\n")

        if args.dry_run:
            conn.rollback()
            print(f"DRY RUN complete — {total} rows validated, transaction ROLLED BACK.")
        else:
            conn.commit()
            print(f"Import complete — {total} rows committed to {args.db}.")

    except Exception as e:
        conn.rollback()
        print(f"\n[ERROR] {e}\nTransaction rolled back.")
        raise
    finally:
        cursor.close()
        conn.close()

    _write_report(report_rows, args.report)


def _write_report(rows: list[dict], path: str) -> None:
    if not rows:
        return
    out = Path(path)
    with open(out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["sheet", "pk", "result", "reason"])
        writer.writeheader()
        writer.writerows(rows)
    print(f"Report written: {out}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser(
        description="Import template_v3.xlsx into SSM_DB (MS SQL Server)"
    )
    p.add_argument("file",         help="Path to template_v3.xlsx")
    p.add_argument("--server",     default="localhost\\SQLEXPRESS",
                   help="SQL Server instance (default: localhost\\SQLEXPRESS)")
    p.add_argument("--db",         default="SSM_DB",
                   help="Database name (default: SSM_DB)")
    p.add_argument("--auth",       choices=["windows", "sql"], default="windows",
                   help="Auth method: windows (default) or sql")
    p.add_argument("--user",       default="",  help="SQL Auth username")
    p.add_argument("--password",   default="",  help="SQL Auth password")
    p.add_argument("--dry-run",    action="store_true",
                   help="Connect + validate + insert in a transaction, then ROLL BACK")
    p.add_argument("--parse-only", action="store_true",
                   help="Read + validate Excel only -- no DB connection needed")
    p.add_argument("--upsert",     action="store_true",
                   help="MERGE: insert new rows, update existing rows (safe to re-run)")
    p.add_argument("--report",     default="import_report.csv",
                   help="Output CSV report path (default: import_report.csv)")
    args = p.parse_args()

    if args.auth == "sql" and not args.user:
        p.error("--user is required when --auth sql")

    run(args)


if __name__ == "__main__":
    main()
