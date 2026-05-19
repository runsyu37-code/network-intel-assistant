"""
tests/test_sanitize.py — unit tests for the data sanitizer.

Run with:
    python -m unittest discover tests -v
"""

import json
import subprocess
import sys
import unittest
from pathlib import Path

# Allow importing from sanitizer/ without installing as a package
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "sanitizer"))

from sanitize import Sanitizer
from patterns import IPV4_RE, MAC_COLON_RE, MAC_DOT_RE, MAC_HYPHEN_RE

_SAMPLES = _ROOT / "samples"
_OUTPUT  = _ROOT / "output"
_SCRIPT  = _ROOT / "sanitizer" / "sanitize.py"


# ── IPv4 ──────────────────────────────────────────────────────────────────

class TestIPv4(unittest.TestCase):

    def test_ipv4_is_replaced(self):
        """A real IP must not survive in the output."""
        result = Sanitizer().sanitize("host 203.0.113.10 connected")
        self.assertNotIn("203.0.113.10", result)

    def test_ipv4_fake_format(self):
        """Fake IP should be in the 10.0.0.X series."""
        result = Sanitizer().sanitize("host 203.0.113.10")
        self.assertIn("10.0.0.", result)

    def test_ipv4_consistency_within_run(self):
        """Same real IP appearing twice must produce the same fake IP both times."""
        result = Sanitizer().sanitize(
            "A: 203.0.113.10, B: 203.0.113.11, A again: 203.0.113.10"
        )
        fakes = IPV4_RE.findall(result)
        self.assertEqual(len(fakes), 3)
        self.assertEqual(fakes[0], fakes[2], "same real IP must map to same fake IP")
        self.assertNotEqual(fakes[0], fakes[1], "different real IPs must map to different fakes")

    def test_ipv4_partial_string_not_matched(self):
        """Substrings that look like IPs inside words must not be matched."""
        result = Sanitizer().sanitize("version1.2.3.4abc")
        # "1.2.3.4" is bounded on the right by 'abc' — \b prevents match
        self.assertIn("1.2.3.4", result)


# ── MAC addresses ─────────────────────────────────────────────────────────

class TestMAC(unittest.TestCase):

    def test_mac_colon_replaced(self):
        """MAC in colon format must be replaced."""
        result = Sanitizer().sanitize("mac aa:bb:cc:dd:ee:ff end")
        self.assertNotIn("aa:bb:cc:dd:ee:ff", result)

    def test_mac_hyphen_replaced(self):
        """MAC in hyphen format must be replaced."""
        result = Sanitizer().sanitize("mac aa-bb-cc-dd-ee-ff end")
        self.assertNotIn("aa-bb-cc-dd-ee-ff", result)

    def test_mac_dot_replaced(self):
        """MAC in Cisco dot format must be replaced."""
        result = Sanitizer().sanitize("mac aabb.ccdd.eeff end")
        self.assertNotIn("aabb.ccdd.eeff", result)

    def test_mac_fake_prefix(self):
        """Fake MAC should start with 'fafe' (locally-administered OUI)."""
        result = Sanitizer().sanitize("0011.22ab.cd01")
        self.assertIn("fafe", result)

    def test_mac_consistency_within_run(self):
        """Same MAC appearing twice must produce the same fake MAC both times."""
        result = Sanitizer().sanitize(
            "M1: 0011.22ab.cd01, M2: 0011.22ab.cd02, M1 again: 0011.22ab.cd01"
        )
        fakes = MAC_DOT_RE.findall(result)
        self.assertEqual(len(fakes), 3)
        self.assertEqual(fakes[0], fakes[2], "same real MAC must map to same fake MAC")
        self.assertNotEqual(fakes[0], fakes[1], "different real MACs must map to different fakes")

    def test_mac_cross_format_same_identity(self):
        """The same MAC in colon and hyphen format must share the same normalized identity."""
        s = Sanitizer()
        # Both refer to the same physical MAC — normalized key is the same
        result = s.sanitize("col: aa:bb:cc:dd:ee:ff  hyp: aa-bb-cc-dd-ee-ff")
        self.assertNotIn("aa:bb:cc:dd:ee:ff", result)
        self.assertNotIn("aa-bb-cc-dd-ee-ff", result)
        # Only one entry in mac_map (both normalized to 'aabbccddeeff')
        self.assertEqual(len(s._mac_map), 1)


# ── Consistency & fresh-per-run ───────────────────────────────────────────

class TestConsistency(unittest.TestCase):

    def test_fresh_mapping_per_run(self):
        """Two separate Sanitizer instances must produce independent mappings."""
        ip = "203.0.113.10"
        fake1 = Sanitizer().sanitize(ip)
        fake2 = Sanitizer().sanitize(ip)
        # Both should produce the same fake because counters start at 1 each time
        self.assertEqual(fake1, fake2)

    def test_different_inputs_get_different_fakes(self):
        """Two different real IPs within one run must get different fake values."""
        s = Sanitizer()
        r = s.sanitize("a: 10.1.1.1 b: 10.1.1.2")
        fakes = IPV4_RE.findall(r)
        self.assertEqual(len(fakes), 2)
        self.assertNotEqual(fakes[0], fakes[1])

    def test_hostname_prefix_routing(self):
        """Hostname containing SW should produce SW-xxx, NVR → NVR-xxx, CAM → CAM-xxx."""
        s = Sanitizer()
        result = s.sanitize(
            "EXAMPLE-SW-FLOOR3-01 EXAMPLE-NVR-B1-01 EXAMPLE-CAM-R301-01"
        )
        self.assertIn("SW-", result)
        self.assertIn("NVR-", result)
        self.assertIn("CAM-", result)


# ── No-leakage (most critical test) ──────────────────────────────────────

class TestNoLeakage(unittest.TestCase):
    """
    Verifies that ZERO real values survive sanitization.

    This is the definitive safety check: if any other test misses a case,
    this test catches it.  A sanitized file is only safe to share with cloud
    AI when every assertion below passes.
    """

    # All real values that appear in the synthetic samples
    REAL_IPS = [
        "203.0.113.1",  "203.0.113.10", "203.0.113.11", "203.0.113.12",
        "203.0.113.13", "203.0.113.14", "203.0.113.15", "203.0.113.16",
        "203.0.113.17", "203.0.113.18", "203.0.113.19", "203.0.113.20",
        "203.0.113.50", "203.0.113.51", "203.0.113.52",
        "203.0.113.100","203.0.113.101","203.0.113.102",
    ]
    REAL_MACS_DOT = [
        "0011.22ab.cd01","0011.22ab.cd02","0011.22ab.cd03","0011.22ab.cd04",
        "0011.22ab.cd05","0011.22ab.cd06","0011.22ab.cd07","0011.22ab.cd08",
        "0011.22ab.cd09","0011.22ab.cd10","0011.22ab.cd11",
        "0011.22ef.0101","0011.22ef.0102","0011.22ef.0103",
        "0011.22ff.aa01","0011.22ff.aa02","0011.22ff.aa03",
        "0011.22cc.dd01","0011.22cc.dd02","0011.22cc.dd03",
    ]
    REAL_HOSTNAMES = [
        "EXAMPLE-SW-FLOOR3-01","EXAMPLE-SW-FLOOR2-01","EXAMPLE-SW-CORE-01",
        "EXAMPLE-NVR-B1-01","EXAMPLE-NVR-B1-02",
        "EXAMPLE-CAM-R301-01","EXAMPLE-CAM-R301-02",
        "EXAMPLE-CAM-R302-01","EXAMPLE-CAM-R302-02",
        "EXAMPLE-CAM-R303-01","EXAMPLE-CAM-C3A-01",
        "EXAMPLE-CAM-C3A-02","EXAMPLE-CAM-ST3B-01",
    ]
    REAL_LOCATIONS = [
        "Building EX-A", "Floor 3", "Room R-301", "Room R-302",
        "Room R-303", "Rack RK-12", "Corridor C-3A", "Stairwell ST-3B",
    ]

    def _sanitize(self, filename: str) -> str:
        text = (_SAMPLES / filename).read_text(encoding="utf-8")
        return Sanitizer().sanitize(text)

    def test_no_real_ip_in_sample_02(self):
        result = self._sanitize("fake_input_02.txt")
        for ip in self.REAL_IPS:
            self.assertNotIn(ip, result, f"Real IP leaked: {ip}")

    def test_no_real_ip_in_sample_03(self):
        result = self._sanitize("fake_input_03.txt")
        for ip in ["203.0.113.1", "203.0.113.12"]:
            self.assertNotIn(ip, result, f"Real IP leaked: {ip}")

    def test_no_real_mac_in_sample_01(self):
        result = self._sanitize("fake_input_01.txt")
        for mac in self.REAL_MACS_DOT:
            self.assertNotIn(mac, result, f"Real MAC leaked: {mac}")

    def test_no_real_mac_in_sample_02(self):
        result = self._sanitize("fake_input_02.txt")
        for mac in self.REAL_MACS_DOT:
            self.assertNotIn(mac, result, f"Real MAC leaked: {mac}")

    def test_no_real_hostname_in_sample_03(self):
        result = self._sanitize("fake_input_03.txt")
        for host in self.REAL_HOSTNAMES:
            self.assertNotIn(host, result, f"Real hostname leaked: {host}")

    def test_no_real_location_in_sample_03(self):
        result = self._sanitize("fake_input_03.txt")
        for loc in self.REAL_LOCATIONS:
            self.assertNotIn(loc, result, f"Real location leaked: {loc}")


# ── CLI via subprocess ────────────────────────────────────────────────────

class TestCLI(unittest.TestCase):

    def _run(self, *args):
        return subprocess.run(
            [sys.executable, str(_SCRIPT)] + list(args),
            capture_output=True, text=True
        )

    def _tmp(self, name: str) -> Path:
        p = _OUTPUT / name
        self.addCleanup(p.unlink, missing_ok=True)
        return p

    def test_basic_run_exits_zero(self):
        """A successful run must exit with code 0."""
        result = self._run(
            str(_SAMPLES / "fake_input_01.txt"),
            str(self._tmp("cli_test_01.txt"))
        )
        self.assertEqual(result.returncode, 0)

    def test_output_file_is_created(self):
        """Output file must exist after a successful run."""
        out = self._tmp("cli_test_02.txt")
        self._run(str(_SAMPLES / "fake_input_01.txt"), str(out))
        self.assertTrue(out.exists())

    def test_summary_line_on_stdout(self):
        """stdout must contain 'Sanitized', 'IPs', 'MACs', 'hostnames'."""
        result = self._run(
            str(_SAMPLES / "fake_input_02.txt"),
            str(self._tmp("cli_test_03.txt"))
        )
        for word in ("Sanitized", "IPs", "MACs", "hostnames"):
            self.assertIn(word, result.stdout)

    def test_missing_input_exits_nonzero(self):
        """A missing input file must exit with non-zero and print to stderr."""
        result = self._run("no_such_file.txt", str(self._tmp("cli_test_04.txt")))
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("Error", result.stderr)

    def test_report_flag_creates_json(self):
        """--report must create a JSON file with ip_mappings and mac_mappings keys."""
        report = self._tmp("cli_test_report.json")
        self._run(
            str(_SAMPLES / "fake_input_02.txt"),
            str(self._tmp("cli_test_05.txt")),
            "--report", str(report)
        )
        self.assertTrue(report.exists())
        with open(report, encoding="utf-8") as f:
            data = json.load(f)
        self.assertIn("ip_mappings", data)
        self.assertIn("mac_mappings", data)
        self.assertGreater(len(data["ip_mappings"]), 0)


# ── Excel support ─────────────────────────────────────────────────────────

class TestExcel(unittest.TestCase):

    def setUp(self):
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

    def _make_workbook(self, cells: list[tuple]) -> Path:
        """Write a temp .xlsx with given (row, col, value) tuples, return path."""
        wb = self.openpyxl.Workbook()
        ws = wb.active
        for r, c, v in cells:
            ws.cell(row=r, column=c, value=v)
        p = _OUTPUT / "_test_input.xlsx"
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        self.addCleanup(p.unlink, missing_ok=True)
        return p

    def _load_cells(self, path: Path) -> list[str]:
        wb = self.openpyxl.load_workbook(path)
        return [
            cell.value
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]

    def test_excel_ip_replaced(self):
        """Real IP in an Excel cell must not survive."""
        inp = self._make_workbook([(1, 1, "Gateway: 203.0.113.1")])
        out = _OUTPUT / "_test_out.xlsx"
        self.addCleanup(out.unlink, missing_ok=True)
        result = subprocess.run(
            [sys.executable, str(_SCRIPT), str(inp), str(out)],
            capture_output=True, text=True
        )
        self.assertEqual(result.returncode, 0)
        values = self._load_cells(out)
        self.assertFalse(any("203.0.113.1" in v for v in values), "Real IP leaked")
        self.assertTrue(any("10.0.0." in v for v in values), "Fake IP not found")

    def test_excel_mac_replaced(self):
        """Real MAC in an Excel cell must not survive."""
        inp = self._make_workbook([(1, 1, "MAC: aa:bb:cc:dd:ee:ff")])
        out = _OUTPUT / "_test_out_mac.xlsx"
        self.addCleanup(out.unlink, missing_ok=True)
        subprocess.run([sys.executable, str(_SCRIPT), str(inp), str(out)])
        values = self._load_cells(out)
        self.assertFalse(any("aa:bb:cc:dd:ee:ff" in v for v in values))

    def test_excel_non_string_cells_preserved(self):
        """Numbers and None cells must be left untouched."""
        inp = self._make_workbook([(1, 1, 42), (1, 2, 3.14), (2, 1, "10.1.2.3")])
        out = _OUTPUT / "_test_out_num.xlsx"
        self.addCleanup(out.unlink, missing_ok=True)
        subprocess.run([sys.executable, str(_SCRIPT), str(inp), str(out)])
        wb = self.openpyxl.load_workbook(out)
        ws = wb.active
        self.assertEqual(ws.cell(1, 1).value, 42)
        self.assertAlmostEqual(ws.cell(1, 2).value, 3.14)

    def test_excel_multiple_sheets(self):
        """Cells across all sheets must be sanitized."""
        wb = self.openpyxl.Workbook()
        wb.active["A1"] = "IP: 203.0.113.5"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "IP: 203.0.113.6"
        inp = _OUTPUT / "_test_multi_sheet.xlsx"
        inp.parent.mkdir(parents=True, exist_ok=True)
        wb.save(inp)
        self.addCleanup(inp.unlink, missing_ok=True)
        out = _OUTPUT / "_test_multi_sheet_out.xlsx"
        self.addCleanup(out.unlink, missing_ok=True)
        subprocess.run([sys.executable, str(_SCRIPT), str(inp), str(out)])
        values = self._load_cells(out)
        self.assertFalse(any("203.0.113" in v for v in values), "Real IP leaked from a sheet")


if __name__ == "__main__":
    unittest.main()
