#!/usr/bin/env python3
"""
SSM Data Import Script
Reads template_v4.xlsx and imports data via the REST API in the correct FK order.

Import order:
  1. Sites  2. Buildings  3. Floors  4. Rooms  5. Racks
  6. NVRs   7. Switches   8. Cameras  9. Users

Usage:
  python import_to_api.py [--file template_v4.xlsx] [--url http://localhost:50680] [--token JWT]
  python import_to_api.py --dry-run          # preview records without posting
"""

import sys
import json
import argparse
import getpass
import openpyxl
import requests

HEADER_ROW = 3
DATA_START_ROW = 4

BOOL_FIELDS = {"has_cctv", "has_nvr", "has_sw", "is_active"}
INT_FIELDS = {
    "floor_count", "floor_number", "total_units", "units_per_u",
    "u_position", "u_subposition", "u_size",
    "total_channels", "active_channels", "retention_days",
    "total_ports", "poe_ports", "poe_budget_w", "poe_used_w",
    "vlan_id", "poe_port_number", "nvr_channel",
}
FLOAT_FIELDS = {"hdd_total_tb", "hdd_used_pct"}

# (sheet_name, api_endpoint, {excel_header: api_field})
# Order matters — FKs must exist before the rows that reference them.
SHEETS = [
    ("1_Site", "/api/sites", {
        "Site_ID (PK)":        "Site_ID",
        "Site Name":           "name",
        "Site Code":           "code",
        "Location / Address":  "location",
        "Description":         "description",
    }),
    ("2_Building", "/api/buildings", {
        "Building_ID (PK)":    "Building_ID",
        "Site_ID (FK)":        "Site_ID",
        "Building Name":       "name",
        "Building Code":       "code",
        "Floor Count":         "floor_count",
        "Description":         "description",
        "Note":                "note",
    }),
    ("3_Floor", "/api/floors", {
        "Floor_ID (PK)":       "Floor_ID",
        "Site_ID (FK)":        "Site_ID",
        "Building_ID (FK)":    "Building_ID",
        "Floor Number":        "floor_number",
        "Floor Name":          "name",
        "Main Function":       "function",
        "Has CCTV?":           "has_cctv",
        "Note":                "note",
    }),
    ("4_Room", "/api/rooms", {
        "Room_ID (PK)":        "Room_ID",
        "Site_ID (FK)":        "Site_ID",
        "Building_ID (FK)":    "Building_ID",
        "Floor_ID (FK)":       "Floor_ID",
        "Room Name":           "name",
        "Room Type":           "type",
        "Has NVR":             "has_nvr",
        "Has SW":              "has_sw",
        "Note":                "note",
    }),
    ("5_Rack", "/api/racks", {
        "Rack_ID (PK)":        "Rack_ID",
        "Site_ID (FK)":        "Site_ID",
        "Building_ID (FK)":    "Building_ID",
        "Floor_ID (FK)":       "Floor_ID",
        "Room_ID (FK)":        "Room_ID",
        "Rack Name":           "name",
        "Total U":             "total_units",
        "Units per U (slots)": "units_per_u",
        "Note":                "note",
    }),
    ("7_NVR", "/api/nvrs", {
        "NVR_ID (PK)":         "NVR_ID",
        "Site_ID (FK)":        "Site_ID",
        "Building_ID (FK)":    "Building_ID",
        "Floor_ID (FK)":       "Floor_ID",
        "Room_ID (FK)":        "Room_ID",
        "Rack_ID (FK)":        "Rack_ID",
        "U Position":          "u_position",
        "U Sub-pos (1-3)":     "u_subposition",
        "U-Size":              "u_size",
        "Device Name":         "device_name",
        "Brand":               "brand",
        "Model":               "model",
        "Serial No (S/N)":     "serial_no",
        "MAC Address":         "mac_address",
        "OS / Firmware":       "os_version",
        "VLAN":                "vlan_id",
        "IP (Internet Port)":  "ip_internet",
        "IP (CCTV Port)":      "ip_cctv",
        "Total Channels":      "total_channels",
        "Active Channels":     "active_channels",
        "HDD Total (TB)":      "hdd_total_tb",
        "Recording Resolution":"recording_res",
        "Retention (days)":    "retention_days",
        "Record Status":       "record_status",
        "Status":              "status",
        "Note":                "notes",
    }),
    ("8_Switch", "/api/poe-switches", {
        "SW_ID(PK)":           "SW_ID",
        "Site_ID (FK)":        "Site_ID",
        "Building_ID (FK)":    "Building_ID",
        "Floor_ID (FK)":       "Floor_ID",
        "Room_ID (FK)":        "Room_ID",
        "Rack_ID (FK)":        "Rack_ID",
        "U Position":          "u_position",
        "U Sub-pos (1-3)":     "u_subposition",
        "Device Name":         "device_name",
        "Switch Type":         "switch_type",
        "Brand":               "brand",
        "Model":               "model",
        "Serial No (S/N)":     "serial_no",
        "MAC Address":         "mac_address",
        "OS / Firmware":       "os_version",
        "VLAN":                "vlan_id",
        "IP Address":          "ip_address",
        "Total Ports":         "total_ports",
        "PoE Ports":           "poe_ports",
        "PoE Budget (W)":      "poe_budget_w",
        "PoE Used (W)":        "poe_used_w",
        "Uplink Port":         "uplink_port",
        "Status":              "status",
        "Note":                "notes",
    }),
    ("6_CCTV", "/api/cameras", {
        "NVR_CH":              "NVR_CH",
        "Site_ID (FK)":        "Site_ID",
        "Building_ID (FK)":    "Building_ID",
        "Floor_ID (FK)":       "Floor_ID",
        "Device Name":         "device_name",
        "Brand":               "brand",
        "Model":               "model",
        "Serial No (S/N)":     "serial_no",
        "MAC Address":         "mac_address",
        "Camera Type":         "camera_type",
        "Resolution":          "resolution",
        "IP Address":          "ip_address",
        "VLAN":                "vlan_id",
        "SW_ID":               "SW_ID",
        "Switch Port":         "poe_port_number",
        "NVR_ID":              "NVR_ID",
        "NVR Channel":         "nvr_channel",
        "install_location":    "install_location",
        "Status":              "status",
        "Note":                "notes",
    }),
    ("9_Users", "/api/users", {
        "Username":            "username",
        "Password":            "password",
        "Display Name":        "display_name",
        "Role":                "role",
        "Is Active":           "is_active",
    }),
]


def coerce(val, field):
    if val is None or str(val).strip() == "":
        return None
    v = str(val).strip()
    if field in BOOL_FIELDS:
        if isinstance(val, bool):
            return val
        return v.lower() in ("yes", "true", "1", "y", "✓")
    if field in INT_FIELDS:
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None
    if field in FLOAT_FIELDS:
        try:
            return float(v)
        except (ValueError, TypeError):
            return None
    return v


def build_col_map(ws, header_map):
    """Map column index → api_field using header row. Uses prefix match as fallback."""
    col_map = {}
    for c in range(1, ws.max_column + 2):
        cell_val = ws.cell(HEADER_ROW, c).value
        if cell_val is None:
            continue
        header = str(cell_val).strip()
        if header in header_map:
            col_map[c] = header_map[header]
            continue
        for key, field in header_map.items():
            if header.startswith(key) or key.startswith(header):
                col_map[c] = field
                break
    return col_map


def read_sheet(ws, col_map):
    records = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        record = {}
        for col_idx, api_field in col_map.items():
            val = row[col_idx - 1] if col_idx <= len(row) else None
            coerced = coerce(val, api_field)
            if coerced is not None:
                record[api_field] = coerced
        if record:
            records.append(record)
    return records


def post_batch(url, records, headers, sheet_name):
    if not records:
        print(f"  [SKIP] {sheet_name}: no data")
        return True
    try:
        resp = requests.post(url, json=records, headers=headers, timeout=30)
    except requests.exceptions.ConnectionError:
        print(f"  [FAIL] {sheet_name}: cannot connect to {url}")
        return False
    if resp.status_code in (200, 201):
        print(f"  [OK]   {sheet_name}: {len(records)} records → HTTP {resp.status_code}")
        return True
    else:
        print(f"  [FAIL] {sheet_name}: HTTP {resp.status_code}")
        print(f"         {resp.text[:300]}")
        return False


def main():
    parser = argparse.ArgumentParser(description="Import SSM Excel template to REST API")
    parser.add_argument("--file",    default="template_v4.xlsx")
    parser.add_argument("--url",     default="http://localhost:50680")
    parser.add_argument("--token",   default="")
    parser.add_argument("--dry-run", action="store_true", help="Preview without posting")
    args = parser.parse_args()

    token = args.token
    if not token and not args.dry_run:
        token = getpass.getpass("Admin JWT token: ").strip()

    auth_headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
    }

    print(f"File : {args.file}")
    print(f"URL  : {args.url}")
    print(f"Mode : {'DRY RUN' if args.dry_run else 'LIVE'}\n")

    wb = openpyxl.load_workbook(args.file, data_only=True)
    all_ok = True

    for sheet_name, endpoint, header_map in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found — skipping")
            continue

        ws = wb[sheet_name]
        col_map = build_col_map(ws, header_map)
        records = read_sheet(ws, col_map)

        if args.dry_run:
            print(f"--- {sheet_name} → {endpoint} ({len(records)} rows)")
            for r in records[:2]:
                print(f"    {json.dumps(r, ensure_ascii=False)}")
            if len(records) > 2:
                print(f"    ... {len(records) - 2} more")
            print()
            continue

        ok = post_batch(args.url.rstrip("/") + endpoint, records, auth_headers, sheet_name)
        if not ok:
            all_ok = False

    if not args.dry_run:
        print("\nAll done." if all_ok else "\nFinished with errors — check output above.")


if __name__ == "__main__":
    main()
